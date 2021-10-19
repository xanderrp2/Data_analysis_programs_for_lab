## Importing librarys and specific moduels 
import os
from openpyxl import load_workbook
from openpyxl import Workbook

## getting root to folder along with files
Files = os.listdir('/Users/ACCOUNT_NAME/Desktop/FOLDER')

## Making lists for things in the future
all_files = []
sheet = []
sheet2 = []
column = []

for i in Files:
    all_files.append("/Users/ACCOUNT_NAME/FOLDER/" + i) ## path directory
for file in all_files:

    ## Creating 2D list to append into excel file (also makes a new one each time)
    row = []

    ## checking to see if this is an excel file
    if sheet2[1] == "xlsx":
        ## Opening excel file
        wb = load_workbook(file,data_only=True)
        worksheet = wb["Sheet1"]
        
        ## Getting values needed for specific calculations from excel file
        B17 = worksheet['B17'].value
        B18 = worksheet['B18'].value
        F21 = worksheet['F21'].value
        F19 = worksheet['F19'].value
        J20 = worksheet['J20'].value
        J19 = worksheet['J19'].value
        B36 = worksheet['B36'].value
        B35 = worksheet['B35'].value
        F38 = worksheet['F38'].value
        F37 = worksheet['F37'].value

        ## making sure that the denominator is not 0 (this makes annoying errors)(also makes the data easier to read)
        if B17 <= 0:
            DF = "N/A"
        else:
            DF = B18/B17
            
        if F19 <= 0:
            AD = "N/A"
        else:
            AD = F21/F19

        if J19 <= 0:
            OC = "N/A"
        else:
            OC = J20/J19

        if B35 <= 0:
            AE = "N/A"
        else: 
            AE = B36/B35

        if F37 <= 0:
            EM = "N/A"
        else:
            EM = F38/F37

        ## Creating worksheet
        WS = Workbook()
        WS["Sheet"].title = "sheet1"
        sheet1 = WS.active

        ## appending values to 2D list
        row.append(sheet2[0])
        row.append(DF)
        row.append(AD)
        row.append(OC)
        row.append(AE)
        row.append(EM)

        ## Appening List to second list (makes it a 2D list)
        column.append(row)

## Appends values to excel file
for i in range(0,len(column),1):
    sheet1.append(column[i])

## Saves all values to excel file
WS.save("EAS data Final.xlsx")
